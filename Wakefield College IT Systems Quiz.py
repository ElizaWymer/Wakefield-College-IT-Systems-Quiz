import math
from openpyxl import load_workbook
import os
import pandas
import random
import threading
import tkinter
import tkinter.messagebox
from tkinter import *
import time 
import winsound

class Login:
    def LoginScreen():
        global usernameInput
        global passwordInput
        global loginButton
        global returnButton
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz Login Menu", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        usernameLabel = tkinter.Label(frame, text = "Username:", font = "calibri 13 bold", fg = textForeground, bg = textBackground,)
        usernameInput = tkinter.Entry(frame, borderwidth = 2, fg = textForeground, bg = textBackground, relief = "ridge", width = 51, font = "calibri 13")
        passwordLabel = tkinter.Label(frame, text = "Password:", font = "calibri 13 bold", fg = textForeground, bg = textBackground,)
        passwordInput = tkinter.Entry(frame, borderwidth = 2, fg = textForeground, bg = textBackground, relief = "ridge", width = 51, font = "calibri 13")
        loginButton = tkinter.Button(frame, text = "Login", font = "calibri 15 bold", command = Login.LoginAuthentification, bg = green, width = 10, height = 3)
        returnButton = tkinter.Button(frame, text = "Return to\n Main Menu", font = "calibri 15 bold", command = Miscellaneous.MainMenu, bg = red, width = 10, height = 3)

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        usernameLabel.grid(column = 0, row = 2, padx = (20,10), pady = 20, sticky = "w")
        usernameInput.focus()
        usernameInput.grid(column = 1, row = 2, sticky = "w")
        passwordLabel.grid(column = 0, row = 3, padx = (20,10), sticky = "w")
        passwordInput.grid(column = 1, row = 3, sticky = "w")
        passwordInput.config(show = "*")
        loginButton.grid(column = 1, row = 5, pady = 25, sticky = "w")
        returnButton.grid(column = 1, row = 5, padx = 355)

    def LoginAuthentification():
        global errorText
        global enteredUsername
        global studentAccount

        enteredUsername = usernameInput.get().lower()
        enteredPassword = passwordInput.get()

        openAccountDetails = pandas.read_excel(spreadsheetName, "Account Details")
        openAccountDataFrame = pandas.DataFrame(openAccountDetails, columns = ["Username", "Password"])

        try:
            Miscellaneous.getAccountType()
            UserIndex = openAccountDataFrame[openAccountDataFrame["Username"] == enteredUsername].index[0]
            cellValue = openAccountDataFrame.iloc[UserIndex]["Username"]
            savedUsername = str(cellValue)
            try:
                cellValue = openAccountDataFrame.iloc[UserIndex]["Password"]
                savedPassword = str(cellValue)
                if accountType == "Admin" and enteredUsername == savedUsername and enteredPassword == savedPassword:
                    studentAccount = False
                    AdminHub.AdminHubScreen()
                elif accountType == "Student" and enteredUsername == savedUsername and enteredPassword == savedPassword:
                    studentAccount = True
                    UserHub.UserHubScreen()
                else:
                    errorText = "You have entered the wrong password."
                    Login.LogInErrorLabel()
                    enteredPassword = None
            except:
                errorText = "You have entered the wrong password."
                Login.LogInErrorLabel()
                enteredPassword = None
        except:
            errorText = "That username does not exist."
            enteredUsername = None
            Login.LogInErrorLabel()

    def LogInErrorLabel():
        errorLabel = tkinter.Label(frame, text = errorText, font = "calibri 13 bold", fg = red, bg = textBackground)
        errorLabel.grid(column = 1, row = 4, sticky = "w")
        usernameInput.config(state = "disable")
        passwordInput.config(state = "disable")
        loginButton.config(state = "disable")
        returnButton.config(state = "disable")
        frame.update()
        time.sleep(1.5)
        errorLabel.destroy()
        usernameInput.config(state = "normal")
        passwordInput.config(state = "normal")
        loginButton.config(state = "normal")
        returnButton.config(state = "normal")

class AdminHub:
    def AdminHubScreen():
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz " + enteredUsername.capitalize() + "'s Admin Configuration Hub", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        takeQuizButton = tkinter.Button(frame, text = "Take\nThe Quiz", font = "calibri 15 bold", command = Quiz.AppendQuestionData, bg = green, width = 15, height = 2)
        createAccountButton = tkinter.Button(frame, text = "Create\nan Account", font = "calibri 15 bold", command = AdminHub.CreateAccountType, bg = green, width = 15, height = 2)
        editAccountButton = tkinter.Button(frame, text = "Edit\nan Account", font = "calibri 15 bold", command = AdminHub.EditAccountEnterAccount, bg = green, width = 15, height = 2)
        deleteAccountButton = tkinter.Button(frame, text = "Delete\nan Account", font = "calibri 15 bold", command = AdminHub.DeleteAccount, bg = red, width = 15, height = 2)
        clearScoreboardButton = tkinter.Button(frame, text = "Scoreboard\nSettings", font = "calibri 15 bold", command = AdminHub.ScoreboardSettings, bg = red, width = 15, height = 2)
        logOutButton = tkinter.Button(frame, text = "Log Out", font = "calibri 15 bold", command = Miscellaneous.MainMenu, bg = red, width = 15, height = 2)
            
        banner.grid(columnspan = 3, row = 1, sticky = "w")
        takeQuizButton.grid(column = 0, row = 2, padx = 5, pady = 40)
        createAccountButton.grid(column = 1, row = 2)
        editAccountButton.grid(column = 2, row = 2)  
        deleteAccountButton.grid(column = 0, row = 3)  
        clearScoreboardButton.grid(column = 1, row = 3)  
        logOutButton.grid(column = 2, row = 3)  

    def ConfirmChange(change):
        Miscellaneous.DestroyFrame()
        confirmBanner = tkinter.Label(frame, text = unitName + " Quiz Confirmation Screen", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        if change == "Create":
            confirmationText = "The account has been successfully created."
        elif change == "TypeStudent":
            confirmationText = "The account has successfully been changed into a student account."            
        elif change == "TypeAdmin":
            confirmationText = "The account has successfully been changed into an admin account."
        elif change == "Username":
            confirmationText = "The account's name has successfully been changed."
        elif change == "Password":
            confirmationText = "The account's password has successfully been changed."
        else:
            workBook = load_workbook(spreadsheetName)
            page = workBook["Account Details"]
            page.delete_rows(userIndex + 2, 1)      
            page = workBook["Account Statistics"]
            page.delete_rows(userIndex + 2, 1)     
            page = workBook["Scoreboard"]
            page.delete_rows(userIndex + 2, 1)     
            workBook.save(filename = spreadsheetName)
            workBook.close()
            confirmationText = "The account has successfully been deleted."
        confirmLable = tkinter.Label(frame, text = confirmationText, font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        returnButton = tkinter.Button(frame, text = "Return to Hub", command = AdminHub.AdminHubScreen, font = "calibri 15 bold", bg = green)

        confirmBanner.grid(columnspan = 3, row = 1)
        confirmLable.grid(columnspan = 3, row = 2, pady = (75, 50))
        returnButton.grid(columnspan = 3, row = 3, pady = 0)

    def CreateAccountType():
        global newAdmin
        
        Miscellaneous.DestroyFrame()
        newAdmin = False
        banner = tkinter.Label(frame, text = unitName + " Quiz Account Creation", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        typeLabel = tkinter.Label(frame, text = "What type of account would you like to create?", font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        adminButton = tkinter.Button(frame, text = "Admin\nAccount", font = "calibri 20 bold", command = AdminHub.CreateAdmin, bg = green, width = 10, height = 4)
        studentButton = tkinter.Button(frame, text = "Student\nAccount", font = "calibri 20 bold", command = AdminHub.CreateAccountMenu, bg = green, width = 10, height = 4)
        cancelButton = tkinter.Button(frame, text = "Return to\nAdmin Hub", font = "calibri 20 bold", command = AdminHub.AdminHubScreen, bg = red, width = 10, height = 4)

        banner.grid(columnspan = 3, row = 1)
        typeLabel.grid(columnspan = 3, row = 2, pady = (25, 15))
        adminButton.grid(column = 0, row = 3, padx = 20)
        studentButton.grid(column = 1, row = 3, sticky = "w")
        cancelButton.grid(column = 2, row = 3, sticky = "w")  

    def CreateAdmin():
        global newAdmin
        newAdmin = True
        AdminHub.CreateAccountMenu()

    def CreateAccountMenu():
        global newUsernameInput
        global newPasswordInput
        global createButton
        global cancelButton

        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz Account Creation", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        newUsernameLabel = tkinter.Label(frame, text = "New Username:", font = "calibri 13 bold", fg = textForeground, bg = textBackground)
        newUsernameInput = tkinter.Entry(frame, borderwidth = 2, fg = textForeground, bg = textBackground, relief = "ridge", width = 47, font = "calibri 13")
        newPasswordLabel = tkinter.Label(frame, text = "New Password:", font = "calibri 13 bold", fg = textForeground, bg = textBackground)
        newPasswordInput = tkinter.Entry(frame, borderwidth = 2, fg = textForeground, bg = textBackground, relief = "ridge", width = 47, font = "calibri 13")
        createButton = tkinter.Button(frame, text = "Create\nAccount", font = "calibri 15 bold", command = AdminHub.CreateAccountAuthentification, bg = green, width = 10, height = 3)
        cancelButton = tkinter.Button(frame, text = "Cancel", font = "calibri 15 bold", command = AdminHub.CreateAccountType, bg = red, width = 10, height = 3)

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        newUsernameLabel.grid(column = 0, row = 2, padx = (20,10), pady = 20, sticky = "w")
        newUsernameInput.grid(column = 1, row = 2, pady = 25, sticky = "w")
        newUsernameInput.focus()
        newPasswordLabel.grid(column = 0, row = 3, padx = (20,10), sticky = "w")
        newPasswordInput.grid(column = 1, row = 3, sticky = "w")
        newPasswordInput.config(show = "*")
        createButton.grid(column = 1, row = 5, pady = 25, sticky = "w")
        cancelButton.grid(column = 1, row = 5, padx = 318)

    def CreateAccountAuthentification():
        global errorText

        if newAdmin == True:
            newType = "Admin"
        else:
            newType = "Student"

        newUsername = newUsernameInput.get()
        newUsername = newUsername.lower()
        newPassword = newPasswordInput.get()
        if len(newUsername) > 0 and len(newUsername) < 15:
            if len(newPassword) > 4 and len(newPassword) < 15:
                if sum(c.isdigit() for c in newPassword) > 2:
                    createaccountDetails = pandas.read_excel(spreadsheetName, "Account Details")
                    createAccountDataFrame = pandas.DataFrame(createaccountDetails, columns = ["Username", "Password"])
                    try:
                        UserIndex = createAccountDataFrame[createAccountDataFrame["Username"] == newUsername].index[0]
                        cellValue = createAccountDataFrame.iloc[UserIndex]["Username"]
                        savedUsername = str(cellValue)
                    except:
                        savedUsername = None
                    try:
                        PassIndex = createAccountDataFrame[createAccountDataFrame["Password"] == newPassword].index[0]
                        cellValue = createAccountDataFrame.iloc[PassIndex]["Password"]
                        savedPassword = str(cellValue)
                    except:
                        savedPassword = None
                    
                    if newUsername != savedUsername:
                        if newPassword != savedPassword:
                            newData = [newType, newUsername, newPassword]

                            workBook = load_workbook(spreadsheetName)
                            page = workBook["Account Details"]
                            page.append(newData)

                            page = workBook["Account Statistics"]
                            newData = [newUsername]

                            columns = 0
                            while columns < 12:
                                newData.append(0)
                                columns += 1

                            page.append(newData)
                            workBook.save(filename = spreadsheetName)
                            workBook.close()
                            
                            AdminHub.ConfirmChange("Create")
                        else:
                            errorText = "An account with that password already exists.\nPlease use a different password."
                            AdminHub.CreateAccountErrorLabel()
                    else:
                        errorText = "An account with that username already exists.\nPlease use a different username."
                        AdminHub.CreateAccountErrorLabel()
                else:
                    errorText = "The password must contain at least three numbers."
                    AdminHub.CreateAccountErrorLabel()
            else:
                errorText = "The password must be between five\nand fifteen characters long."
                AdminHub.CreateAccountErrorLabel()
        else:
            errorText = "The account name must be between one\nand fifteen characters long."
            AdminHub.CreateAccountErrorLabel()

    def CreateAccountErrorLabel():
        errorLabel = tkinter.Label(frame, text = errorText, font = "calibri 13 bold", fg = red, bg = textBackground)
        errorLabel.grid(column = 1, row = 4, sticky = "w")
        newUsernameInput.config(state = "disable")
        newPasswordInput.config(state = "disable")
        createButton.config(state = "disable")
        cancelButton.config(state = "disable")
        frame.update()
        time.sleep(1.5)
        errorLabel.destroy()
        newUsernameInput.config(state = "normal")
        newPasswordInput.config(state = "normal")
        createButton.config(state = "normal")
        cancelButton.config(state = "normal")

    def EditAccountEnterAccount():
        global editAccountInput
        global editConfirm
        global editCancel
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz Account Editing", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        editLabel = tkinter.Label(frame, text = "Enter the name of the account you want to edit", font = "Calibri 15 bold", fg = textForeground, bg = textBackground)
        editAccountInput = tkinter.Entry(frame, borderwidth = 2, fg = textForeground, bg = textBackground, relief = "ridge", width = 47, font = "calibri 13")
        editConfirm = tkinter.Button(frame, text = "Submit", font = "calibri 15 bold", command = AdminHub.EditAccountAuthentification, bg = green)
        editCancel = tkinter.Button(frame, text = "Cancel", font = "calibri 15 bold", command = AdminHub.AdminHubScreen, bg = red)

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        editLabel.grid(columnspan = 3, row = 2, pady = 35)
        editAccountInput.grid(columnspan = 3, row = 3)
        editAccountInput.focus()
        editConfirm.grid(column = 0, row = 5, padx = (59, 225), pady = 35)
        editCancel.grid(column = 1, row = 5)

    def EditAccountAuthentification():
        global changingAccount
        global errorLabel

        changingAccount = editAccountInput.get()
                
        accountDetails = pandas.read_excel(spreadsheetName, "Account Details")
        accountDataFrame = pandas.DataFrame(accountDetails, columns = ["Username"])

        num = 0
        for x in accountDataFrame["Username"]:
            x = accountDataFrame.iloc[num]["Username"]
            usernames.append(x)
            num += 1

            for x in usernames:
                if x.lower() == changingAccount.lower():
                    accountFound = True
                    break
                else:
                    accountFound = False

        usernames.clear()

        if accountFound == True:
            AdminHub.EditAccountTypeOrUserOrPass()
        else:
            errorLabel = tkinter.Label(frame, text = "The username you\nentered does not exist", font = "calibri 15 bold", fg = red, bg = textBackground)
            errorLabel.grid(columnspan = 3, row = 4, pady = (15, 0))
            AdminHub.EditAccountErrorLabel()

    def EditAccountErrorLabel():
        editAccountInput.config(state = "disable")
        editConfirm.config(state = "disable")
        editCancel.config(state = "disable")
        frame.update()
        time.sleep(1.5)
        errorLabel.destroy()
        editAccountInput.config(state = "normal")
        editConfirm.config(state = "normal")
        editCancel.config(state = "normal")

    def EditAccountTypeOrUserOrPass():
        global editBanner
        global editLabel
        editBanner = None
        editLabel = None
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz Account Editing", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        editAccountUserOrPassLabel = tkinter.Label(frame, text = "Would you like to change the\naccount's priviliges, username or password?", font = "Calibri 15 bold", fg = textForeground, bg = textBackground)
        privButton = tkinter.Button(frame, text = "Priviliges", font = "calibri 15 bold", command = AdminHub.EditPriviliges, bg = green, width = 10, height = 2)
        usernameButton = tkinter.Button(frame, text = "Username", font = "calibri 15 bold", command = AdminHub.EditUsername, bg = green, width = 10, height = 2)
        passwordButton = tkinter.Button(frame, text = "Password", font = "calibri 15 bold", command = AdminHub.EditPassword, bg = green, width = 10, height = 2)
        cancelButton = tkinter.Button(frame, text = "Cancel", font = "calibri 15 bold", command = AdminHub.EditAccountEnterAccount, bg = red, width = 10, height = 2)
        
        banner.grid(columnspan = 3, row = 1, sticky = "w")
        editAccountUserOrPassLabel.grid(columnspan = 3, row = 2, pady = 25)
        privButton.grid(column = 0, row = 3, padx = (40, 0), pady = 20, sticky = "w")
        usernameButton.grid(column = 0, row = 3, padx = (170, 0), sticky = "w")
        passwordButton.grid(column = 0, row = 3, padx = (310, 0), sticky = "w")
        cancelButton.grid(column = 0, row = 3, padx = (450, 0), sticky = "w")

    def EditPriviliges():
        global newType
        Miscellaneous.DestroyFrame()
        accountDetails = pandas.read_excel(spreadsheetName, "Account Details")
        accountDataFrame = pandas.DataFrame(accountDetails, columns = ["Username", "Account Type"])
        userIndex = accountDataFrame[accountDataFrame["Username"] == changingAccount.lower()].index[0] #This line is tempermental.
        cellValue = accountDataFrame.iloc[userIndex]["Account Type"]
        accountType = str(cellValue)

        if accountType == "Admin":
            newType = "a student"
        else:
            newType = "an admin"

        editbanner = tkinter.Label(frame, text = unitName + " Quiz Privilige Editing", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        editLabel = tkinter.Label(frame, text = "Are you sure you want to change\n" + changingAccount + "'s account to " + newType + " account?", font = "Calibri 15 bold", fg = textForeground, bg = textBackground)
        confirmButton = tkinter.Button(frame, text = "Yes", command = AdminHub.ChangeAccountType, bg = green, font = "calibri 15 bold")
        cancelButton = tkinter.Button(frame, text = "No", command = AdminHub.EditAccountTypeOrUserOrPass, bg = red, font = "calibri 15 bold")

        editbanner.grid(columnspan = 3, row = 1, sticky = "w")
        editLabel.grid(columnspan = 3, row = 2, pady = 25)
        confirmButton.grid(column = 0, row = 3)
        cancelButton.grid(column = 1, row = 3) 

    def ChangeAccountType():
        if len(newType) == 9:
            test = newType[2:].capitalize()
        else:
            test = newType[3:].capitalize()
        workBook = load_workbook(spreadsheetName)
        page = workBook["Account Details"]
        accountDetails = pandas.read_excel(spreadsheetName, "Account Details")
        accountDataFrame = pandas.DataFrame(accountDetails, columns = ["Username"])
        accountDetailsIndex = accountDataFrame[accountDataFrame["Username"] == changingAccount].index[0]
        page.cell(row=accountDetailsIndex + 2, column= 1).value = test
        workBook.save(filename = spreadsheetName)
        workBook.close()
        if test == "Student":
            AdminHub.ConfirmChange("TypeStudent")
        else:
            AdminHub.ConfirmChange("TypeAdmin")

    def EditUsername():
        AdminHub.EditAccountTypeOrUserOrPass.editBanner = "Username"
        AdminHub.EditAccountTypeOrUserOrPass.editLabel = "username"
        AdminHub.EditAccountDetails()

    def EditPassword():
        AdminHub.EditAccountTypeOrUserOrPass.editBanner = "Password"
        AdminHub.EditAccountTypeOrUserOrPass.editLabel = "password"
        AdminHub.EditAccountDetails()

    def EditAccountDetails():
        global editDetailsInput
        global editConfirm
        global editCancel
        Miscellaneous.DestroyFrame()
        editbanner = tkinter.Label(frame, text = unitName + " Quiz " + AdminHub.EditAccountTypeOrUserOrPass.editBanner + " Editing", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        editLabel = tkinter.Label(frame, text = "Please enter the account's new " + AdminHub.EditAccountTypeOrUserOrPass.editLabel, font = "Calibri 15 bold", fg = textForeground, bg = textBackground)
        editDetailsInput = tkinter.Entry(frame, borderwidth = 2, fg = textForeground, bg = textBackground, relief = "ridge", width = 47, font = "calibri 13")
        editConfirm = tkinter.Button(frame, text = "Submit", font = "calibri 15 bold", command = AdminHub.EditAccountDetailsAuthentification, bg = green)
        editCancel = tkinter.Button(frame, text = "Cancel", font = "calibri 15 bold", command = AdminHub.EditAccountTypeOrUserOrPass, bg = red)

        editbanner.grid(columnspan = 3, row = 1, sticky = "w")
        editLabel.grid(columnspan = 3, row = 2, pady = 25)
        editDetailsInput.grid(columnspan = 3, row = 3)
        editDetailsInput.focus()
        if AdminHub.EditAccountTypeOrUserOrPass.editBanner == "Password":
            editDetailsInput.config(show = "*")
        editConfirm.grid(column = 0, row = 5, padx = (59, 225), pady = 25)
        editCancel.grid(column = 1, row = 5)

    def EditAccountDetailsAuthentification():
        global errorLabel
        newDetail = editDetailsInput.get()
        accountDetails = pandas.read_excel(spreadsheetName, "Account Details")
        accountDataFrame = pandas.DataFrame(accountDetails, columns = ["Username"])
        if AdminHub.EditAccountTypeOrUserOrPass.editBanner == "Username":
            if len(newDetail) > 0 and len(newDetail) < 15:
                try:
                    workBook = load_workbook(spreadsheetName)
                    page = workBook["Account Details"]
                    accountDetailsIndex = accountDataFrame[accountDataFrame["Username"] == changingAccount].index[0]
                    page.cell(row=accountDetailsIndex + 2, column= 2).value = newDetail
                    workBook.save(filename = spreadsheetName)
                    workBook.close()
                    AdminHub.ConfirmChange("Username")
                except:
                    errorLabel = tkinter.Label(frame, text = "An account with that username already exists.\nPlease use a different username.", font = "calibri 13 bold", fg = red, bg = textBackground)
                    errorLabel.grid(columnspan = 3, row = 4, padx = (135, 0), pady = (25, 0), sticky = "w")
                    AdminHub.EditDetailsErrorLabel()
            else:
                errorLabel = tkinter.Label(frame, text = "The account name must be between one\nand fifteen characters long.", font = "calibri 13 bold", fg = red, bg = textBackground)
                errorLabel.grid(columnspan = 3, row = 4, padx = (150, 0), pady = (25, 0), sticky = "w")
                AdminHub.EditDetailsErrorLabel()
        else:
            if len(newDetail) > 4 and len(newDetail) < 15:
                if sum(c.isdigit() for c in newDetail) > 2:
                    workBook = load_workbook(spreadsheetName)
                    page = workBook["Account Details"]
                    accountDetailsIndex = accountDataFrame[accountDataFrame["Username"] == changingAccount].index[0]
                    page.cell(row=accountDetailsIndex + 2, column= 3).value = newDetail
                    workBook.save(filename = spreadsheetName)
                    workBook.close()
                    AdminHub.ConfirmChange("Password")
                else:
                    errorLabel = tkinter.Label(frame, text = "The password must contain at least three numbers.", font = "calibri 13 bold", fg = red, bg = textBackground)
                    errorLabel.grid(columnspan = 3, row = 4, padx = (130, 0), pady = (25, 0), sticky = "w")
                    AdminHub.EditDetailsErrorLabel()  
            else:
                errorLabel = tkinter.Label(frame, text = "The password must be between five\nand fifteen characters long.", font = "calibri 13 bold", fg = red, bg = textBackground)
                errorLabel.grid(columnspan = 3, row = 4, padx = (160, 0), pady = (25, 0), sticky = "w")
                AdminHub.EditDetailsErrorLabel()

    def EditDetailsErrorLabel():
        editDetailsInput.config(state = "disable")
        editConfirm.config(state = "disable")
        editCancel.config(state = "disable")
        frame.update()
        time.sleep(1.5)
        errorLabel.destroy()
        editDetailsInput.config(state = "normal")
        editConfirm.config(state = "normal")
        editCancel.config(state = "normal")

    def DeleteAccount():
        global deleteInput
        global deleteConfirm
        global deleteCancel
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz Account Deleting", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        deleteLabel = tkinter.Label(frame, text = "Enter the name of the account you want to delete", font = "Calibri 15 bold", fg = textForeground, bg = textBackground)
        deleteInput = tkinter.Entry(frame, borderwidth = 2, fg = textForeground, bg = textBackground, relief = "ridge", width = 47, font = "calibri 13")
        deleteConfirm = tkinter.Button(frame, text = "Submit", font = "calibri 15 bold", command = AdminHub.DeleteAuthentification, bg = green)
        deleteCancel = tkinter.Button(frame, text = "Cancel", font = "calibri 15 bold", command = AdminHub.AdminHubScreen, bg = red)

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        deleteLabel.grid(columnspan = 3, row = 2, pady = 25)
        deleteInput.grid(columnspan = 3, row = 3)
        deleteInput.focus()
        deleteConfirm.grid(column = 0, row = 5, padx = (57, 227), pady = 25)
        deleteCancel.grid(column = 1, row = 5)

    def DeleteAuthentification():
        global errorLabel
        global username
        global userIndex
        global types

        username = deleteInput.get()
        if username == enteredUsername:
            errorLabel = tkinter.Label(frame, text = "You cannot delete\nthe account you're currently using.", font = "calibri 13 bold", fg = red, bg = textBackground)
            errorLabel.grid(columnspan = 3, row = 4, padx = (220, 0), pady = (25, 0), sticky = "w")
            AdminHub.DeleteAccountErrorLabel()
        else:
            deleteAccount = pandas.read_excel(spreadsheetName, "Account Details")
            deleteAccountDataFrame = pandas.DataFrame(deleteAccount, columns = ["Account Type"])
            try:
                deleteAccount = pandas.read_excel(spreadsheetName, "Account Details")
                deleteAccountDataFrame = pandas.DataFrame(deleteAccount, columns = ["Username", "Password"])
                userIndex = deleteAccountDataFrame[deleteAccountDataFrame["Username"] == username].index[0]
                AdminHub.DeleteConfirm()
            except:
                errorLabel = tkinter.Label(frame, text = "The username you\nentered does not exist.", font = "calibri 13 bold", fg = red, bg = textBackground)
                errorLabel.grid(columnspan = 3, row = 4, padx = (220, 0), pady = (25, 0), sticky = "w")
                AdminHub.DeleteAccountErrorLabel()

    def DeleteAccountErrorLabel():
        deleteInput.config(state = "disable")
        deleteConfirm.config(state = "disable")
        deleteCancel.config(state = "disable")
        frame.update()
        time.sleep(1.5)
        errorLabel.destroy()
        deleteInput.config(state = "normal")
        deleteConfirm.config(state = "normal")
        deleteCancel.config(state = "normal")

    def DeleteConfirm():
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz Account Deleting", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        deleteAccountLabel = tkinter.Label(frame, text = "Are you sure that you want to delete this account?", font = "Calibri 15 bold", fg = textForeground, bg = textBackground)
        deleteConfirmButton = tkinter.Button(frame, text = "Yes", font = "calibri 20 bold", command = lambda: AdminHub.ConfirmChange("Delete"), bg = red, width = 10, height = 4)
        deleteCancelButton = tkinter.Button(frame, text = "No", font = "calibri 20 bold", command = AdminHub.DeleteAccount, bg = green, width = 10, height = 4)

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        deleteAccountLabel.grid(columnspan = 3, row = 2, pady = 25)
        deleteConfirmButton.grid(column = 1, row = 3, sticky = "w")
        deleteCancelButton.grid(column = 2, row = 3, sticky = "w")

    def ScoreboardSettings():
        pass

class UserHub:
    def UserHubScreen():
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz " + enteredUsername.capitalize() + "'s Student Hub", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        takeQuizButton = tkinter.Button(frame, text = "Take The\nQuiz", font = "calibri 15 bold", command = Quiz.AppendQuestionData, bg = green, width = 25, height = 3)
        userStatsButton = tkinter.Button(frame, text = "Your Stats", font = "calibri 15 bold", command = UserHub.StatsPage1, bg = green, width = 25, height = 3)
        viewScoreboardButton = tkinter.Button(frame, text = "View\nScoreboard", font = "calibri 15 bold", command = Scoreboard.SetupScoreboard, bg = green, width = 25, height = 3)
        logOutButton = tkinter.Button(frame, text = "Log Out", font = "calibri 15 bold", command = Miscellaneous.MainMenu, bg = red, width = 25, height = 3)

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        takeQuizButton.grid(column = 0, row = 2, padx = (23, 0), pady = 25)
        userStatsButton.grid(column = 1, row = 2)  
        viewScoreboardButton.grid(column = 0, row = 3, padx = (23, 0))  
        logOutButton.grid(column = 1, row = 3)  

    def StatsPage1():
        global page1
        global page2
        global page3

        page1 = True
        page2 = False
        page3 = False
        UserHub.DisplayUserStats()

    def StatsPage2():
        global page1
        global page2
        global page3
        
        page1 = False
        page2 = True
        page3 = False
        UserHub.DisplayUserStats()

    def StatsPage3():
        global page1
        global page2
        global page3
        page1 = False
        page2 = False
        page3 = True
        UserHub.DisplayUserStats()

    def DisplayUserStats():
        Miscellaneous.DestroyFrame()      

        userStatsDetails = pandas.read_excel(spreadsheetName, "Account Statistics")
        statsDataFrame = pandas.DataFrame(userStatsDetails, columns = [
                                                                       "Username", "Number of Quizzes Completed", 
                                                                       "Total Number of Statements Given", "Total Number of Statements Filled in Correctly",
                                                                       "Percentage of Statements Filled in Correctly", "Total Number of Skips Used", 
                                                                       "Highest Score", "Average Score", "Fastest Time with Highest Score", "Average Time (Minutes:Seconds)"
                                                                      ])

        userIndex = statsDataFrame[statsDataFrame["Username"] == enteredUsername].index[0]
        userStatsList = list(statsDataFrame.iloc[userIndex])

        statList = [str(x) for x in userStatsList]

        if page1 == True:

            banner = tkinter.Label(frame, text = unitName + " Quiz " + statList[0].capitalize() + "'s Stats Page 1", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
            statLabel1 = tkinter.Label(frame, text = "Number of Quizzes Completed:                                            " + statList[1], font = "calibri 15 bold", fg = textForeground, bg = textBackground)
            statLabel2 = tkinter.Label(frame, text = "\nHighest Score:                                                                         " + statList[6], font = "calibri 15 bold", fg = textForeground, bg = textBackground)
            statLabel3 = tkinter.Label(frame, text = "\nAverage Score:                                                                        " + statList[7], font = "calibri 15 bold", fg = textForeground, bg = textBackground)        
            previousPage = UserHub.StatsPage1
            nextPage = UserHub.StatsPage2
        if page2 == True:
            banner = tkinter.Label(frame, text = unitName + " Quiz " + statList[0].capitalize() + "'s Stats Page 2", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
            statLabel1 = tkinter.Label(frame, text = "Total Number of Statements Given:                                     " + statList[2], font = "calibri 15 bold", fg = textForeground, bg = textBackground)
            statLabel2 = tkinter.Label(frame, text = "\nTotal Number of Statements Filled in Correctly:                 " + statList[3], font = "calibri 15 bold", fg = textForeground, bg = textBackground)
            statLabel3 = tkinter.Label(frame, text = "\nPercentage of Statements Filled in Correctly:                   " + statList[4], font = "calibri 15 bold", fg = textForeground, bg = textBackground)                            
            previousPage = UserHub.StatsPage1
            nextPage = UserHub.StatsPage3
        elif page3 == True:
            banner = tkinter.Label(frame, text = unitName + " Quiz " + statList[0].capitalize() + "'s Stats Page 3", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
            statLabel1 = tkinter.Label(frame, text = "Total Number of Skips Used:                                                 " + statList[5], font = "calibri 15 bold", fg = textForeground, bg = textBackground)
            statLabel2 = tkinter.Label(frame, text = "\nFastest Time with Highest Score:                                     " + statList[8], font = "calibri 15 bold", fg = textForeground, bg = textBackground)
            statLabel3 = tkinter.Label(frame, text = "\nAverage Time Taken to Complete a Quiz:                       " + statList[9], font = "calibri 15 bold", fg = textForeground, bg = textBackground) 
            previousPage = UserHub.StatsPage2
            nextPage = UserHub.StatsPage3

        prevButton = tkinter.Button(frame, text = "Previous Page", font = "calibri 15 bold", command = previousPage, bg = green)
        returnButton = tkinter.Button(frame, text = "Return to Your Hub", font = "calibri 15 bold", command = UserHub.UserHubScreen, bg = red)
        nextButton = tkinter.Button(frame, text = "  Next Page  ", font = "calibri 15 bold", command = nextPage, bg = green)

        if page1 == True:
            prevButton.config(state = "disable")
        elif page3 == True:
            nextButton.config(state = "disable")

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        statLabel1.grid(column = 0, row = 2, padx = (50, 0), pady = (25,0), sticky = "w")
        statLabel2.grid(column = 0, row = 3, padx = (50, 0), sticky = "w")
        statLabel3.grid(column = 0, row = 4, padx = (50, 0), pady = (0,25), sticky = "w")
        prevButton.grid(column = 0, row = 5, padx = (50, 0), sticky = "w")
        returnButton.grid(columnspan = 1, row = 5, padx = (220, 595))
        nextButton.grid(columnspan = 1, row = 5)

class Quiz:
    def AppendQuestionData():
        global sections
        global questions
        global answers

        global questionNumber
        global usedSkips
        global score
        global answeredCorrectly

        sections = []
        questions = []
        answers = []

        num = 0
        for x in SQAHDataFrame["Sections"]:
            x = SQAHDataFrame.iloc[num]["Sections"]
            sections.append(x)
            num += 1

        num = 0
        for x in SQAHDataFrame["Questions"]:
            x = SQAHDataFrame.iloc[num]["Questions"]
            questions.append(x)
            num += 1

        num = 0
        for x in SQAHDataFrame["Answers"]:
            x = SQAHDataFrame.iloc[num]["Answers"]
            answers.append(x)
            num += 1

        sections = [str(x) for x in sections]
        badEggs = []
        num = 0

        while num != len(sections):
            if sections[num] == "nan":
                badEggs.append(sections[num])
            num += 1

        for x in badEggs:
            sections.remove(x)

        questions = [str(x) for x in questions]
        badEggs = []
        num = 0

        while num != len(questions):
            if questions[num] == "nan":
                badEggs.append(questions[num])
            num += 1

        for x in badEggs:
            questions.remove(x)
        
        answers = [str(x) for x in answers]
        badEggs = []
        num = 0

        while num != len(answers):
            if answers[num] == "nan":
                badEggs.append(answers[num])
            num += 1

        for x in badEggs:
            answers.remove(x)

        timer = threading.Thread(name="background", target = Quiz.Timer)
        timer.start()

        questionNumber = 0
        usedSkips = 0
        score = 0
        answeredCorrectly = 0

        Quiz.GenerateQuestion()

    def GenerateQuestion():
        global newSection
        global newQuestion
        global newAnswer
        global newHint
        global tries
        global questionNumber
        global totalTries

        #Sound the hard coding alarm!!! Fixing this would be a whole different can of worms that I dare not touch.
        selectSection = random.randint(0,6)
        if sections[selectSection] == "Section A - Digital Devices in IT Systems":
            selectQuestion = random.randint(0,4)
        elif sections[selectSection] == "Section B - Data Transmission":
            selectQuestion = random.randint(5,9)
        elif sections[selectSection] == "Section C - Issues with Data Transmission":
            selectQuestion = random.randint(10,14)
        elif sections[selectSection] == "Section D - Operating Online":
            selectQuestion = random.randint(15,19)
        elif sections[selectSection] == "Section E - Protecting Data and Information":
            selectQuestion = random.randint(20,24)
        elif sections[selectSection] == "Section F - Impacts of IT Systems":
            selectQuestion = random.randint(25,29)
        elif sections[selectSection] == "Section G - Issues Caused by IT Systems":
            selectQuestion = random.randint(30,34)
        totalTries = systemConfigDataFrame.iloc[1]["Data"]
        tries = totalTries
        newSection = sections[selectSection]
        try:
            newQuestion = questions[selectQuestion].replace("\\n", "\n")
            newAnswer = answers[selectQuestion]
            subAnswer = newAnswer[1:len(newAnswer)]
        
            splitAnswer = newAnswer.split()
            newHint = ""
            for x in splitAnswer:
                subAnswer = x[1:len(x)]
                subHint = x[0]
                for x in subAnswer:
                    subHint += "-"
                newHint += subHint + " "

            newQuestion = newQuestion.replace("REPLACE", newHint)

            del questions[selectQuestion]
            del answers[selectQuestion]
            questionNumber += 1
        except:
            Quiz.GenerateQuestion()

        Quiz.QuizScreen()

    def Timer():
        global rawTime
        global quizEnd
        rawTime = 0
        quizEnd = False
        while True:
            rawTime += 1
            time.sleep(1)
            if quizEnd == True:
                break

    def QuizScreen():
        global answerInput
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz: Statement " + str(questionNumber), font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        answerInput = tkinter.Entry(frame, fg = textForeground, bg = textBackground, relief = "ridge", font = "calibri 15", width = 30)
        answerSubmitButton = tkinter.Button(frame, text = "Submit Answer", font = "calibri 15 bold", command = Quiz.SubmitAnswer, bg = green)
        answerSkipButton = tkinter.Button(frame, text = "Skip Statement", font = "calibri 15 bold", command = lambda: Quiz.AnswerResults("Skipped"), bg = red)

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        answerInput.grid(columnspan = 3, row = 4, pady = 25)
        answerInput.focus()
        answerSubmitButton.grid(column = 0, row = 5, padx = 60)
        answerSkipButton.grid(column = 1, row = 5)

        totalSkips = systemConfigDataFrame.iloc[2]["Data"]

        if usedSkips == totalSkips or questionNumber == 10:
            answerSkipButton.config(state = "disable")

        Quiz.DisplayQuestion()

    def DisplayQuestion():
        topicLabel = tkinter.Label(frame, text = newSection, font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        questionLabel = tkinter.Label(frame, text = newQuestion, font = "calibri 15", fg = textForeground, bg = textBackground)

        topicLabel.grid(columnspan = 3, row = 2, pady = 15)
        questionLabel.grid(columnspan = 3, row = 3)

    def SubmitAnswer():
        global tries
        global score
        global answeredCorrectly
        userAnswer = answerInput.get()
        if userAnswer.lower() == newAnswer.lower():
            if tries == totalTries:
                score += 3
            elif tries < totalTries and tries > 0:
                score += 1
            answeredCorrectly += 1
            Quiz.AnswerResults("Correct!")
        else:
            tries -= 1
            Quiz.AnswerResults("Incorrect!")

    def AnswerResults(results):
        global quizEnd
        global usedSkips
        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz: Statement " + str(questionNumber) + " Results", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        if results == "Correct!" and tries == totalTries and questionNumber != 10:
            resultsLabel = tkinter.Label(frame, text = "\n" + results + "!\nYou get three points.\nTotal Points: " + str(score), font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        elif results == "Correct!" and tries < totalTries and tries > 0 and questionNumber != 10:
            resultsLabel = tkinter.Label(frame, text = "\n" + results + "!\nYou get one point.\nTotal Points: " + str(score), font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        elif results == "Correct!" and questionNumber == 10:
            resultsLabel = tkinter.Label(frame, text = "\n" + results + "\nYou filled in all of the statements correctly!\nWell done!", font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        elif results == "Incorrect!" and tries < totalTries and tries > 0 or results == "Incorrect" and questionNumber == 11 and tries < totalTries and tries > 0:
            resultsLabel = tkinter.Label(frame, text = "\n" + results + "\nTry again.\nYou have " + str(tries) + " try/tries left.", font = "calibri 15 bold", fg = textForeground, bg = textBackground)
            resultsButton = tkinter.Button(frame, text = "Try\nAgain", font = "calibri 15 bold", command = Quiz.QuizScreen, bg = green)
        elif results == "Incorrect!" and tries == 0:
            resultsLabel = tkinter.Label(frame, text = "\n" + results + "\nThe answer was: " + newAnswer + "\n\nYou failed to fill in the statement correctly twice!\nGame over!", font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        elif results == "Skipped":
            usedSkips += 1
            questionLabel = tkinter.Label(frame, text = newQuestion, font = "calibri 15", fg = textForeground, bg = textBackground)
            resultsLabel = tkinter.Label(frame, text = "\nThe answer was: " + newAnswer + "\nSkips Used: " + str(usedSkips) + "/3", font = "calibri 15 bold", fg = textForeground, bg = textBackground)
            questionLabel.grid(columnspan = 3, row = 2, pady = (15,0))

        if results == "Correct!" and questionNumber != 10 or results == "Skipped":
            resultsButton = tkinter.Button(frame, text = "Next\nStatement", command = Quiz.GenerateQuestion, font = "calibri 15 bold", bg = green)
        elif results == "Incorrect!" and tries == 0 or results == "Correct!" and questionNumber == 10:
            quizEnd = True
            resultsButton = tkinter.Button(frame, text = "View\nResults", font = "calibri 15 bold", command = Quiz.RecordUserStats, bg = green)
        
        banner.grid(columnspan = 3, row = 1, sticky = "w")
        resultsLabel.grid(columnspan = 3, row = 3)
        resultsButton.grid(columnspan = 3, row = 4, pady = 15)

    def ConvertTime(time):
        global quizTime
        global hub

        minutes = []
        splitTime = str(time / 60).rsplit(".")
        for x in splitTime:
            minutes.append(x)
        seconds = round(int(minutes[1]) / (10**len(minutes[1]))  * 60)

        if int(minutes[0]) < 10:
            quizTime = "0" + str(minutes[0])
        else:
            quizTime = str(minutes[0])
        if seconds < 10:
            quizTime += ":0" + str(seconds)
        else:   
            quizTime += ":" + str(seconds)

        if studentAccount == True:
            hub = UserHub.UserHubScreen
        else:
            hub = AdminHub.AdminHubScreen

        return quizTime

    def RecordUserStats():
        userStatsDetails = pandas.read_excel(spreadsheetName, "Account Statistics")
        statsDataFrame = pandas.DataFrame(userStatsDetails, columns = [
                                                                       "Username", "Number of Quizzes Completed", 
                                                                       "Total Number of Statements Given", "Total Number of Statements Filled in Correctly",
                                                                       "Percentage of Statements Filled in Correctly", "Total Number of Skips Used", 
                                                                       "Highest Score","Total Score", "Average Score", "Fastest Time with Highest Score (Seconds)",
                                                                       "Fastest Time with Highest Score", "Total number of seconds", 
                                                                       "Average Time (Seconds)", "Average Time (Minutes:Seconds)"
                                                                      ])

        userIndex = statsDataFrame[statsDataFrame["Username"] == enteredUsername].index[0]
        userStatsList = list(statsDataFrame.iloc[userIndex])

        userStatsList[1] += 1
        userStatsList[2] += questionNumber
        userStatsList[3] += answeredCorrectly

        percentageOfQuestions = userStatsList[3] / userStatsList[2]
        userStatsList[4] = str(round((percentageOfQuestions * 100))) + "%"
        userStatsList[5] += usedSkips

        if score > userStatsList[6]:
             userStatsList[6] = score     
             userStatsList[9] = rawTime
             userStatsList[10] = Quiz.ConvertTime(userStatsList[9])

        userStatsList[7] += score
        userStatsList[8] = round(userStatsList[7] / userStatsList[1])

        if score == userStatsList[6] and rawTime < userStatsList[9]:
            userStatsList[9] = rawTime
            userStatsList[10] = Quiz.ConvertTime(userStatsList[9])
        
        userStatsList[11] += rawTime
        userStatsList[12] = userStatsList[11] / userStatsList[1]
        userStatsList[13] = Quiz.ConvertTime(userStatsList[12])

        workBook = load_workbook(spreadsheetName)
        page = workBook["Account Statistics"]
        indexStart = 0
        for x in userStatsList:
            page.cell(row=userIndex + 2, column = userStatsList.index(x, indexStart) + 1).value = x
            indexStart += 1

        workBook.save(filename = spreadsheetName)
        workBook.close()

        Quiz.OverallResults()

    def OverallResults():
        global soundFileName

        Miscellaneous.DestroyFrame()
        soundEffectsEnabled = systemConfigDataFrame.iloc[3]["Bool"]
        if soundEffectsEnabled == 1: 
            if score == 0:
                respone = "Congratulations, you completley failed!"
                soundFileName = "Failure.wav"
            elif score > 0 and score < 12:
                respone = "You did okay. Better luck next time."
                soundFileName = "Cheer.wav"
            elif score >= 12 and score < 21:
                respone = "You did quite well. Well done."
                soundFileName = "Small Applause.wav"   
            elif score >= 21 and score < 30:
                respone = "You got an amazing score, congratulations!"
                soundFileName = "Applause.wav"   
            else:
                respone = "You got a perfect score. I'm impressed."
                soundFileName = "Victory.wav"

        banner = tkinter.Label(frame, text = unitName + " Quiz: Quiz Results", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        scoreLabel = tkinter.Label(frame, text = "Score: " + str(score), font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        percentageLabel = tkinter.Label(frame, text = "Percentage of Score: " + str(int(score /30 * 100)) + "%" , font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        
        questions.clear()
        answers.clear()

        finalTime = Quiz.ConvertTime(rawTime)

        timeLabel = tkinter.Label(frame, text = "Time: " + finalTime, font = "calibri 15 bold", fg = textForeground, bg = textBackground)      
        resultsResponse = tkinter.Label(frame, text = respone, font = "calibri 15 bold", fg = textForeground, bg = textBackground)
        returnButton = tkinter.Button(frame, text = "Return to Your Hub", command = hub, bg = green, font = "calibri 15 bold")
        
        banner.grid(columnspan = 3, row = 1, sticky = "w")
        scoreLabel.grid(columnspan = 3, row = 2, pady = (25,0))
        percentageLabel.grid(columnspan = 3, row = 3)
        timeLabel.grid(columnspan = 3, row = 4)
        resultsResponse.grid(columnspan = 3, row = 5, pady = (0,25))
        returnButton.grid(columnspan = 3, row = 6)
        endmusic = threading.Thread(name='background', target= Quiz.EndMusic)
        endmusic.start()

    def EndMusic():
        winsound.PlaySound(soundPath + soundFileName, winsound.SND_FILENAME)

class Scoreboard:
    def SetupScoreboard():
        global usernames
        global scores
        global rawTimes
        global displayTimes
        global numberOfPages
        global page
        global values
        
        usernames = []
        scores = []
        rawTimes = []
        displayTimes = []

        spreadsheetName = "Wakefield College IT Systems Quiz Data.xlsx"
        scoreboardDetails = pandas.read_excel(spreadsheetName, "Scoreboard")
        scoreboardDataFrame = pandas.DataFrame(scoreboardDetails, columns = ["Username", "Score", "Raw Time", "Display Time"])

        for x in scoreboardDataFrame["Username"]:
            usernames.append(str(x))

        for x in scoreboardDataFrame["Score"]:
            scores.append(int(x))

        for x in scoreboardDataFrame["Raw Time"]:
            rawTimes.append(int(x))

        for x in scoreboardDataFrame["Display Time"]:
            displayTimes.append(str(x))

        usernames = [x for _, x in sorted(zip(scores, usernames), reverse = True)]
        rawTimes = [x for _, x in sorted(zip(scores, rawTimes), reverse = True)]
        displayTimes = [x for _, x in sorted(zip(scores, displayTimes), reverse = True)]
        scores = sorted(scores, reverse = True)

        test = 0.499

        oc_set = set()
        res = []
        for idx, val in enumerate(scores):
            if val not in oc_set:
                for x in res:
                    try:
                        if rawTimes[x] < rawTimes[y]:
                            scores[y] -= test
                        elif rawTimes[y] < rawTimes[x]:
                            scores[x] -= test
                        test -= 0.001
                    except:
                        pass
                    y = x
                x = None
                y = None
                oc_set.add(val)
                res.clear()         
            res.append(idx)

        usernames = [x for _, x in sorted(zip(scores, usernames), reverse = True)]
        rawTimes = [x for _, x in sorted(zip(scores, rawTimes), reverse = True)]
        displayTimes = [x for _, x in sorted(zip(scores, displayTimes), reverse = True)]
        scores = sorted(scores, reverse = True)
        for x in scores:
            scores[scores.index(x)] = round(x)

        numberOfPages = math.ceil(len(usernames) / 4)
        page = 1
        values = [0, 1, 2, 3]
        Scoreboard.ScoreboardStructure()

    def PreviousPage():
        global page
        global values
        Miscellaneous.DestroyFrame()
        page -= 1
        for x in values:
            values[values.index(x)] = x - 4
        Scoreboard.ScoreboardStructure()

    def NextPage():
        global page
        global values
        Miscellaneous.DestroyFrame()
        page += 1
        for x in values:
            values[values.index(x)] = x + 4
        Scoreboard.ScoreboardStructure()

    def ScoreboardStructure():
        Miscellaneous.DestroyFrame()
        try:
            usernameHeader = tkinter.Label(frame, text = " Username ", font = "calibri 15 bold", bg = blue, width = 20, borderwidth = 1, relief = "solid")
            usernameHeader.grid(column = 1, row = 2, sticky = "w", padx = (65,0), pady = (25,0))
            scoreHeader = tkinter.Label(frame, text = " High Score ", font = "calibri 15 bold", bg = blue, width = 10, borderwidth = 1, relief = "solid")
            scoreHeader.grid(column = 1, row = 2, sticky = "w", padx = (269, 0), pady = (25,0))
            timeHeader = tkinter.Label(frame, text = " Time ", font = "calibri 15 bold", bg = blue, width = 10, borderwidth = 1, relief = "solid")
            timeHeader.grid(column = 1, row = 2, sticky = "w", padx = (373,0), pady = (25,0))

            firstLine = tkinter.Label(frame, text = usernames[values[0]], font = "calibri 15 bold", width = 20, borderwidth = 1, relief = "solid") 
            firstLine.grid(column = 1, row = 3, padx = (65,0), sticky = "w")
            sLine = tkinter.Label(frame, text = scores[values[0]], font = "calibri 15 bold", width = 10, borderwidth = 1, relief = "solid")
            sLine.grid(column = 1, row = 3, sticky = "w", padx = (269,0))
            tLine = tkinter.Label(frame, text = displayTimes[values[0]], font = "calibri 15 bold", width = 10, borderwidth = 1, relief = "solid")
            tLine.grid(column = 1, row = 3, sticky = "w", padx = (373,0))

            tomLine1 = tkinter.Label(frame, text = usernames[values[1]], font = "calibri 15 bold", width = 20, borderwidth = 1, relief = "solid") 
            tomLine1.grid(column = 1, row = 4, padx = (65,0), sticky = "w")
            tomLine2 = tkinter.Label(frame, text = scores[values[1]], font = "calibri 15 bold", width = 10, borderwidth = 1, relief = "solid")
            tomLine2.grid(column = 1, row = 4, sticky = "w", padx = (269,0))
            tomLine3 = tkinter.Label(frame, text = displayTimes[values[1]], font = "calibri 15 bold", width = 10, borderwidth = 1, relief = "solid")
            tomLine3.grid(column = 1, row = 4, sticky = "w", padx = (373,0))

            pLine = tkinter.Label(frame, text = usernames[values[2]], font = "calibri 15 bold", width = 20, borderwidth = 1, relief = "solid") 
            pLine.grid(column = 1, row = 5, padx = (65,0), sticky = "w")
            pLine2 = tkinter.Label(frame, text = scores[values[2]], font = "calibri 15 bold", width = 10, borderwidth = 1, relief = "solid")
            pLine2.grid(column = 1, row = 5, sticky = "w", padx = (269,0))
            pLine3 = tkinter.Label(frame, text = displayTimes[values[2]], font = "calibri 15 bold", width = 10, borderwidth = 1, relief = "solid")
            pLine3.grid(column = 1, row = 5, sticky = "w", padx = (373,0))

            aLine = tkinter.Label(frame, text = usernames[values[3]], font = "calibri 15 bold", width = 20, borderwidth = 1, relief = "solid") 
            aLine.grid(column = 1, row = 6, padx = (65,0), sticky = "w")
            aLine2 = tkinter.Label(frame, text = scores[values[3]], font = "calibri 15 bold", width = 10, borderwidth = 1, relief = "solid")
            aLine2.grid(column = 1, row = 6, sticky = "w", padx = (269,0))
            aLine3 = tkinter.Label(frame, text = displayTimes[values[3]], font = "calibri 15 bold", width = 10, borderwidth = 1, relief = "solid")
            aLine3.grid(column = 1, row = 6, sticky = "w", padx = (373,0))
        except:
            pass

        banner = tkinter.Label(frame, text = unitName + " Quiz Scoreboard Page " + str(page), font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        banner.grid(columnspan = 4, row = 1, sticky = "w")      

        prevButton = tkinter.Button(frame, text = "Previous\nPage", font = "calibri 15 bold", command = Scoreboard.PreviousPage, bg = green, width = 10)
        prevButton.grid(column = 1, row = 8, sticky = "w", padx = (64, 0), pady = 12)

        if studentAccount == True:
            hub = UserHub.UserHubScreen
        else:
            hub = AdminHub.AdminHubScreen

        hubButton = tkinter.Button(frame, text = "Return to\nYour Hub", font = "calibri 15 bold", command = hub, bg = red)
        hubButton.grid(column = 1, row = 8, sticky = "w", padx = (223, 0))

        nextButton = tkinter.Button(frame, text = "Next\nPage", font = "calibri 15 bold", command = Scoreboard.NextPage, bg = green, width = 10)
        nextButton.grid(column = 1, row = 8, sticky = "w", padx = (366, 0))

        if page == 1:
            prevButton.config(state = "disable")
        elif page == numberOfPages:
            nextButton.config(state = "disable")

class Miscellaneous:
    def MainMenu():
        global banner
        global loginButton
        global closeButton
        if colourTheme == "Light":
            themeButtonText = "Enable\nDark Mode"
        else:
            themeButtonText = "Enable\nLight Mode"

        Miscellaneous.DestroyFrame()
        banner = tkinter.Label(frame, text = unitName + " Quiz Main Menu", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
        loginButton = tkinter.Button(frame, text = "Login", font = "calibri 15 bold", command = Login.LoginScreen, bg = green, width = 25, height = 3) 
        rulesButton = tkinter.Button(frame, text = "View\nRules", font = "calibri 15 bold", command = Miscellaneous.RulesPage1, bg = green, width = 25, height = 3)
        themeButton = tkinter.Button(frame, text = themeButtonText, font = "calibri 15 bold", command = Miscellaneous.ChangeTheme, bg = blue, width = 25, height = 3)        
        closeButton = tkinter.Button(frame, text = "Quit", font = "calibri 15 bold", command = Miscellaneous.CloseButton, bg = red, width = 25, height = 3)

        banner.grid(columnspan = 3, row = 1)
        loginButton.grid(column = 0, row = 2, padx = 20, pady = 25)
        rulesButton.grid(column = 1, row = 2, sticky = "w")
        themeButton.grid(column = 0, row = 3)
        closeButton.grid(column = 1, row = 3, sticky = "w")  

    def RulesPage1():
        global page1
        global page2
        global page3

        page1 = True
        page2 = False
        page3 = False
        Miscellaneous.RulesGrid()

    def RulesPage2():
        global page1
        global page2
        global page3
        
        page1 = False
        page2 = True
        page3 = False
        Miscellaneous.RulesGrid()

    def RulesPage3():
        global page1
        global page2
        global page3
        page1 = False
        page2 = False
        page3 = True
        Miscellaneous.RulesGrid()

    def RulesGrid():
        Miscellaneous.DestroyFrame()       
        if page1 == True:
            banner = tkinter.Label(frame, text = unitName + " Quiz Rules Page 1", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
            ruleLabel = tkinter.Label(
                                      frame, text = "This quiz is designed to test your knowledge of the\n" + unitName + " unit while providing effective exam answer structure."
                                                    "\n\nYou will be given ten statements, each with a missing word.\nYou will then have two chances to fill in the missing word correctly.",
                                                     font = "calibri 15 bold", fg = textForeground, bg = textBackground
                                     )        
            previousPage = Miscellaneous.RulesPage1
            nextPage = Miscellaneous.RulesPage2
        if page2 == True:
            banner = tkinter.Label(frame, text = unitName + " Quiz Rules Page 2", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
            ruleLabel = tkinter.Label(
                                        frame, text = "If the statement is completed correctly\non the first try you will be given three points."
                                                      "\n\nIf the statement is completed correctly\non the second try you will be given one point.",
                                        font = "calibri 15 bold", fg = textForeground, bg = textBackground
                                    )
            previousPage = Miscellaneous.RulesPage1
            nextPage = Miscellaneous.RulesPage3
        elif page3 == True:
            banner = tkinter.Label(frame, text = unitName + " Quiz Rules Page 3", font = "calibri 18 bold", bg = blue, width = 46, height = 1)
            ruleLabel = tkinter.Label(
                                        frame, text = "If the statement is not completed\ncorrectly within two tries the quiz ends."
                                                      "\n\nStatements can be skipped but\nyou can only use three skips per quiz.",
                                        font = "calibri 15 bold", fg = textForeground, bg = textBackground
                                    )
            previousPage = Miscellaneous.RulesPage2
            nextPage = Miscellaneous.RulesPage3

        prevButton = tkinter.Button(frame, text = "Previous Page", font = "calibri 15 bold", command = previousPage, bg = green)
        returnButton = tkinter.Button(frame, text = "Return to Main Menu", font = "calibri 15 bold", command = Miscellaneous.MainMenu, bg = red)
        nextButton = tkinter.Button(frame, text = "Next Page", font = "calibri 15 bold", command = nextPage, bg = green)

        if page1 == True:
            prevButton.config(state = "disable")
        elif page3 == True:
            nextButton.config(state = "disable")

        banner.grid(columnspan = 3, row = 1, sticky = "w")
        ruleLabel.grid(columnspan = 3, row = 2, pady = 25)
        prevButton.grid(column = 0, row = 3)
        returnButton.grid(column = 1, row = 3)
        nextButton.grid(column = 2, row = 3)

    def ChangeTheme():
        global colourTheme
        global textForeground
        global textBackground
        global red 
        global blue 
        global green

        if colourTheme == "Light":
            colourTheme = "Dark"
            window.configure(bg = "#2b373d")
            frame.configure(bg = "#2b373d")
            textForeground = "white"
            textBackground = "#2b373d"
            red = "#B90103"
            green = "forest green"
            blue = "#0784b5"
        else:
            colourTheme = "Light"
            window.configure(bg = "#f0f0f0")
            frame.configure(bg = "#f0f0f0")
            textForeground = "black"
            textBackground = "#f0f0f0"
            red = "#f19c99"
            green = "#97d077"
            blue = "#99ccff"
        Miscellaneous.MainMenu()

    def getAccountType():
        global UserIndex
        global accountType
        accountDetails = pandas.read_excel(spreadsheetName, "Account Details")
        accountDataFrame = pandas.DataFrame(accountDetails, columns = ["Account Type", "Username"])
        UserIndex = accountDataFrame[accountDataFrame["Username"] == enteredUsername].index[0]
        cellValue = accountDataFrame.iloc[UserIndex]["Account Type"]
        accountType = str(cellValue)

    def CloseButton():
        if tkinter.messagebox.askyesno("Warning", "Are you sure you want to close the application?", icon = "warning") == True:
            window.destroy()

    def DestroyFrame():
        try:
            for widgets in frame.winfo_children():
                widgets.destroy()
        except:
            pass

#The file path for the sound effects 
soundPath = os.getcwd() + os.path.join("\\SoundEffects\\")

#Data Frames
spreadsheetName = "Wakefield College IT Systems Quiz Data.xlsx"
systemConfigDetails = pandas.read_excel(spreadsheetName, "System Configuration")
systemConfigDataFrame = pandas.DataFrame(systemConfigDetails, columns = ["Data", "Bool"])
accountDetails = pandas.read_excel(spreadsheetName, "Account Details")
accountDataFrame = pandas.DataFrame(accountDetails, columns = ["Account Type", "Username", "Password"])
SQAHDetails = pandas.read_excel(spreadsheetName, "Sections, Questions and Answers")
SQAHDataFrame = pandas.DataFrame(SQAHDetails, columns = ["Sections", "Questions", "Answers"])

usernames = []

unitName = systemConfigDataFrame.iloc[0]["Data"]

if not systemConfigDataFrame.iloc[4]["Bool"]:
    colourTheme = "Dark"
else:
    colourTheme = "Light"

#Variables to set window
window = tkinter.Tk()
window.geometry("605x288")
window.resizable(width = False, height = False)
window.title("Wakefield College " + unitName + " Quiz")
frame = Frame(window)
frame.grid()

Miscellaneous.ChangeTheme()
window.mainloop()
