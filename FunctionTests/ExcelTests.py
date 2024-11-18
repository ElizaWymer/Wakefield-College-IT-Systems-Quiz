import pandas as pd
import random
import openpyxl 
from openpyxl import load_workbook


def ReadAccountData():
    inputName = input("Enter your username: ")
    inputPass = input("Enter your password: ")

    try:
        accountDetailsIndex = accountDataFrame[accountDataFrame["Username"] == inputName].index[0]
        cellValue = accountDataFrame.iloc[accountDetailsIndex]["Username"]
        username = str(cellValue)
        try:
            cellValue = accountDataFrame.iloc[accountDetailsIndex]["Password"]
            password = str(cellValue)
        except:
            print("You have entered the wrong password.")
            password = ""
    except:
        print("That username does not exist.")
        username = ""

    if inputName == username and inputPass == password:
        print("Success")
    else:
        print("Fail")


def ReadSQAH():

    sections = []
    questions = []
    answers = []
    
    num = 0
    for x in SQAHDataFrame["Sections"]:
        x = SQAHDataFrame.iloc[num]["Sections"]
        sections.append(x)
        num += 1

    num = 0
    for x in SQAHDataFrame["Questions"]:
        x = SQAHDataFrame.iloc[num]["Questions"]
        questions.append(x)
        num += 1

    num = 0
    for x in SQAHDataFrame["Answers"]:
        x = SQAHDataFrame.iloc[num]["Answers"]
        answers.append(x)
        num += 1

    sections = [str(x) for x in sections]
    badEggs = []
    num = 0

    while num != len(sections):
        if sections[num] == "nan":
            badEggs.append(sections[num])
        num += 1

    for x in badEggs:
        sections.remove(x)

    questions = [str(x) for x in questions]
    badEggs = []
    num = 0

    while num != len(questions):
        if questions[num] == "nan":
            badEggs.append(questions[num])
        num += 1

    for x in badEggs:
        questions.remove(x)
    
    answers = [str(x) for x in answers]
    badEggs = []
    num = 0

    while num != len(answers):
        if answers[num] == "nan":
            badEggs.append(answers[num])
        num += 1

    for x in badEggs:
        answers.remove(x)

    selectSection = random.randint(0,6)

    if sections[selectSection] == "Section A - Digital Devices in IT Systems":
        selectQuestion = random.randint(0,4)
    elif sections[selectSection] == "Section B - Data Transmission":
        selectQuestion = random.randint(5,9)
    elif sections[selectSection] == "Section C - Issues with Data Transmission":
        selectQuestion = random.randint(10,14)
    elif sections[selectSection] == "Section D - Operating Online":
        selectQuestion = random.randint(15,19)
    elif sections[selectSection] == "Section E - Protecting Data and Information":
        selectQuestion = random.randint(20,24)
    elif sections[selectSection] == "Section F - Impacts of IT Systems":
        selectQuestion = random.randint(25,29)
    elif sections[selectSection] == "Section G - Issues Caused by IT Systems":
        selectQuestion = random.randint(30,34)

    print(sections[selectSection])
    print(questions[selectQuestion])
    print(answers[selectQuestion])

    sections.remove(sections[selectSection])
    questions.remove(questions[selectQuestion])
    answers.remove(answers[selectQuestion])

def WriteData():
    accType = input("Type:")
    accName = input("Name:")
    accPass = input("Pass:")
    newData = [accType, accName, accPass] 

    workBook = load_workbook(spreadName)
    page = workBook.active
    page.append(newData)
    
    workBook.save(filename = spreadName)
    workBook.close()

def EditData():
    workBook = load_workbook(spreadName)
    page = workBook.active
    newType = "Student"
    inputName = "Elizabeth"
    accountDetailsIndex = accountDataFrame[accountDataFrame["Username"] == inputName].index[0]
    print(accountDetailsIndex)
    page.cell(row=accountDetailsIndex + 2, column= 1).value = newType
    workBook.save(filename = spreadName)
    workBook.close()

def Delete():
    deleteAccount = pd.read_excel(spreadName, "Account Details")
    deleteAccountDataFrame = pd.DataFrame(deleteAccount, columns = ["Account Type"])
    types = deleteAccountDataFrame["Account Type"].values.tolist()
    print(types.count("Admin"))

spreadName = "Wakefield College IT Systems Quiz Data.xlsx"

accountDetails = pd.read_excel(spreadName, "Account Details")
accountDataFrame = pd.DataFrame(accountDetails, columns = ["Account Type", "Username", "Password"])
SQAHDetails = pd.read_excel(spreadName, "Sections, Questions and Answers")
SQAHDataFrame = pd.DataFrame(SQAHDetails, columns = ["Sections", "Questions", "Answers", "Hints"])



ReadSQAH()
